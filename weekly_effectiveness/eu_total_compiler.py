import pprint

import openpyxl
from openpyxl.utils import column_index_from_string
import json
import os
from constants import COUNTRIES
import dotenv
from pathlib import Path

env_path = Path('.', '.env')
dotenv.load_dotenv(dotenv_path=env_path)


def fill_in_total_dict_for_geo(chanel, geo_dict, total_row_num, sheet):
    global columns
    columns_to_be_filled_in = ['C', 'D', 'F', 'G', 'H', 'I', 'J', 'L', 'M', 'N', 'O', 'P', 'Q']
    if chanel == 'traff':
        titles_row = 2
        columns = columns_to_be_filled_in + ['B', 'E', 'K']
    elif chanel == 'base':
        titles_row = geo_dict['traff']['total_row'] + 4
        columns = columns_to_be_filled_in
    elif chanel == 'mp':
        titles_row = geo_dict['base']['total_row'] + 4
        columns = columns_to_be_filled_in
    for column in columns:
        column_title = sheet.cell(row=titles_row, column=column_index_from_string(column)).value
        total_value = sheet.cell(row=total_row_num, column=column_index_from_string(column)).value
        geo_dict[chanel][column_title] = total_value


class EUTotalCompiler:
    """Class reads data from the formulas file and fills in EU_total and saves all togather as a new file"""
    def __init__(self, json_file_path):
        self.json_file_path = json_file_path
        self.formulas_file_path: str = None
        self.formulas_file_open = None
        self.eu_total_sheet = None
        self.geos_total_dict = {}  # Stores data from total rows for geos
        self.start_date = None
        self.end_date = None
        self.traffic_start_row = 1
        self.traffic_geo_row = self.traffic_start_row + 2
        self.traffic_end_row = 24
        self.base_start_row = 27
        self.base_geo_row = self.base_start_row + 2
        self.base_end_row = 59
        self.mp_start_row = 62
        self.mp_geo_row = self.mp_start_row + 2
        self.mp_end_row = 94

    def get_info_from_json(self):
        with open(self.json_file_path) as json_file:
            data_passed = json.load(json_file)
        self.start_date = data_passed['start_date']
        self.end_date = data_passed['end_date']
        self.formulas_file_path = data_passed['formulas_file_path']

        # Creating dicts for each chanel inside each geo, and adding info on the total row number.
        for geo in data_passed.keys():
            if geo not in ('start_date', 'end_date', 'formulas_file_path'):
                geo_data = data_passed.get(geo)
                self.geos_total_dict.setdefault(geo, {})
                for chanel in geo_data.keys():
                    self.geos_total_dict[geo][chanel] = {'total_row': geo_data[chanel]}

    def get_data_from_formulas_file(self):
        self.formulas_file_open = openpyxl.load_workbook(self.formulas_file_path, data_only=True)

        # Iterting through sheets and collecting data from total rows
        for sheet_name in self.formulas_file_open.sheetnames:
            if sheet_name != 'EU_total':
                sheet_geo = sheet_name
                if sheet_geo not in ('mx', 'pe'):
                    traff_total_row = self.geos_total_dict.get(sheet_geo, {}).get('traff', {}).get('total_row')
                    base_total_row = self.geos_total_dict.get(sheet_geo, {}).get('base', {}).get('total_row')
                    mp_total_row = self.geos_total_dict.get(sheet_geo, {}).get('mp', {}).get('total_row')
                    if traff_total_row:
                        fill_in_total_dict_for_geo(chanel='traff', geo_dict=self.geos_total_dict[sheet_geo],
                                                   total_row_num=traff_total_row, sheet=self.formulas_file_open[sheet_name])
                    if base_total_row:
                        fill_in_total_dict_for_geo(chanel='base', geo_dict=self.geos_total_dict[sheet_geo],
                                                   total_row_num=base_total_row, sheet=self.formulas_file_open[sheet_name])
                    if mp_total_row:
                        fill_in_total_dict_for_geo(chanel='mp', geo_dict=self.geos_total_dict[sheet_geo],
                                                   total_row_num=mp_total_row, sheet=self.formulas_file_open[sheet_name])

    def fill_in_eu_total_table(self):
        global geo_row
        self.eu_total_sheet = self.formulas_file_open['EU_total']
        for total_geo in self.geos_total_dict.keys():
            if total_geo != 'mx':
                geo_total_info_dict = self.geos_total_dict.get(total_geo)
                for chanel in geo_total_info_dict.keys():
                    if chanel == 'traff':
                        geo_row = self.traffic_geo_row
                    elif chanel == 'base':
                        geo_row = self.base_geo_row
                    elif chanel == 'mp':
                        geo_row = self.mp_geo_row
                    self.eu_total_sheet.cell(row=geo_row, column=column_index_from_string('A')).value = \
                        COUNTRIES.get(total_geo)
                    self.eu_total_sheet.cell(row=geo_row, column=column_index_from_string('C')).value = \
                        geo_total_info_dict[chanel]['Sell, pcs']
                    self.eu_total_sheet.cell(row=geo_row, column=column_index_from_string('D')).value = \
                        geo_total_info_dict[chanel]['Sell, EUR']
                    self.eu_total_sheet.cell(row=geo_row, column=column_index_from_string('F')).value = \
                        geo_total_info_dict[chanel]['Bought, pcs']
                    self.eu_total_sheet.cell(row=geo_row, column=column_index_from_string('G')).value = \
                        geo_total_info_dict[chanel]['Income']
                    self.eu_total_sheet.cell(row=geo_row, column=column_index_from_string('H')).value = \
                        geo_total_info_dict[chanel]['% bought']
                    self.eu_total_sheet.cell(row=geo_row, column=column_index_from_string('I')).value = \
                        geo_total_info_dict[chanel]['VAT']
                    self.eu_total_sheet.cell(row=geo_row, column=column_index_from_string('J')).value = \
                        geo_total_info_dict[chanel]['Courier fee']
                    self.eu_total_sheet.cell(row=geo_row, column=column_index_from_string('L')).value = \
                        geo_total_info_dict[chanel]['Production']
                    self.eu_total_sheet.cell(row=geo_row, column=column_index_from_string('M')).value = \
                        geo_total_info_dict[chanel]['OPEX']
                    self.eu_total_sheet.cell(row=geo_row, column=column_index_from_string('N')).value = \
                        geo_total_info_dict[chanel]['Costs']
                    self.eu_total_sheet.cell(row=geo_row, column=column_index_from_string('O')).value = \
                        geo_total_info_dict[chanel]['Profit, EUR']
                    self.eu_total_sheet.cell(row=geo_row, column=column_index_from_string('P')).value = \
                        geo_total_info_dict[chanel]['Margin, %']
                    self.eu_total_sheet.cell(row=geo_row, column=column_index_from_string('Q')).value = \
                        geo_total_info_dict[chanel]['Avg check']
                    if chanel == 'traff':
                        self.eu_total_sheet.cell(row=geo_row, column=column_index_from_string('B')).value = \
                            geo_total_info_dict[chanel]['Leads, pcs']
                        self.eu_total_sheet.cell(row=geo_row, column=column_index_from_string('E')).value = \
                            geo_total_info_dict[chanel]['% approved']
                        self.eu_total_sheet.cell(row=geo_row, column=column_index_from_string('K')).value = \
                            geo_total_info_dict[chanel]['Advertising']

                    # Updating current row
                    geo_row += 1
                    if chanel == 'traff':
                        self.traffic_geo_row = geo_row
                    elif chanel == 'base':
                        self.base_geo_row = geo_row
                    elif chanel == 'mp':
                        self.mp_geo_row = geo_row

    def filling_headers_and_totals(self):
        # removing empty lines
        self.eu_total_sheet.delete_rows(self.traffic_geo_row, self.traffic_end_row - self.traffic_geo_row)
        traffic_rows_removed = self.traffic_end_row - self.traffic_geo_row
        self.traffic_end_row -= self.traffic_end_row - self.traffic_geo_row
        self.base_start_row -= traffic_rows_removed
        self.base_geo_row -= traffic_rows_removed
        self.base_end_row -= traffic_rows_removed
        self.mp_start_row -= traffic_rows_removed
        self.mp_geo_row -= traffic_rows_removed
        self.mp_end_row -= traffic_rows_removed

        self.eu_total_sheet.delete_rows(self.base_geo_row, self.base_end_row - self.base_geo_row)
        base_rows_removed = self.base_end_row - self.base_geo_row
        self.base_end_row -= self.base_end_row - self.base_geo_row
        self.mp_start_row -= base_rows_removed
        self.mp_geo_row -= base_rows_removed
        self.mp_end_row -= base_rows_removed

        self.eu_total_sheet.delete_rows(self.mp_geo_row, self.mp_end_row - self.mp_geo_row)
        self.mp_end_row -= self.mp_end_row - self.mp_geo_row

        # Filling headers
        for row in (self.traffic_start_row, self.base_start_row, self.mp_start_row):
            self.eu_total_sheet.cell(row=row, column=column_index_from_string('B')).value = \
                f'{self.start_date} - {self.end_date}'

        # Filling totals
        for start_row, last_data_row, end_row in (
                (self.traffic_start_row, self.traffic_geo_row, self.traffic_end_row),
                (self.base_start_row, self.base_geo_row, self.base_end_row),
                (self.mp_start_row, self.mp_geo_row, self.mp_end_row)
        ):
            first_data_row = start_row + 2
            if start_row == 1:  # check if we are working with traffic table
                self.eu_total_sheet.cell(row=end_row, column=column_index_from_string('B')).value = \
                    f'=SUM(B{first_data_row}:B{last_data_row - 1})'
                self.eu_total_sheet.cell(row=end_row, column=column_index_from_string('E')).value = \
                    f'=(C{end_row}/B{end_row})'
                self.eu_total_sheet.cell(row=end_row, column=column_index_from_string('K')).value = \
                    f'=SUM(K{first_data_row}:K{last_data_row - 1})'
            for sum_collumn_letter in ('C', 'D', 'F', 'G', 'I', 'J', 'L', 'M', 'N', 'O'):
                self.eu_total_sheet.cell(row=end_row, column=column_index_from_string(sum_collumn_letter)).value = \
                    f'=SUM({sum_collumn_letter}{first_data_row}:{sum_collumn_letter}{last_data_row - 1})'
            self.eu_total_sheet.cell(row=end_row, column=column_index_from_string('H')).value = \
                f'=F{end_row}/C{end_row}'
            self.eu_total_sheet.cell(row=end_row, column=column_index_from_string('P')).value = \
                f'=O{end_row}/D{end_row}'
            self.eu_total_sheet.cell(row=end_row, column=column_index_from_string('Q')).value = \
                f'=G{end_row}/F{end_row}'

    def saving_as_new_file(self):
        self.formulas_file_open.save(os.path.join(os.path.dirname(self.formulas_file_path),
                                                  f'eff_{self.start_date}_{self.end_date}_total.xlsx'))

    def run(self):
        self.get_info_from_json()
        self.get_data_from_formulas_file()
        self.fill_in_eu_total_table()
        self.filling_headers_and_totals()
        self.saving_as_new_file()
