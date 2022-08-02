import datetime
import os

from openpyxl.styles import Border, Side
import requests
import openpyxl
from openpyxl.utils import column_index_from_string
from bs4 import BeautifulSoup
from pprint import pprint
import dotenv
from pathlib import Path

env_path = Path('.', '.env')
dotenv.load_dotenv(dotenv_path=env_path)

template_address = input('Type in template address\n')
save_file_address = input('Type in file address for complete report\n')

RATES = {}  # Rates information in format {'web name'(str): {'geo name' (str): (web tariff(int), network tariff(int))}}


class GeoDictGenerator:

    def __init__(self):
        self.geos_dict = dict()
        self.webs_dict = dict()
        self.start_date = ''
        self.end_date = ''
        self.session = requests.Session()
        self.web_codes = {}  # Web codes in format {'web_code'(str): 'web name' (str)}
        self.country_rows = dict()

    def compilation_geos_and_webs_lists(self, start_date, end_date):
        self.session.post(os.getenv('platform_login_link'),
                          data={'in_user': os.getenv('platform_username'), 'in_pass': os.getenv('platform_password')})
        self.start_date = start_date
        self.end_date = end_date
        general_info_response = self.session.get(
            f'{os.getenv("platform_calls_analytics_link")}?from={start_date}&to={end_date}&o=&c=&geo=')
        soup_general_info = BeautifulSoup(general_info_response.text, 'lxml')
        tbody_general_info = soup_general_info.find_all('tbody')
        web_flag = False
        offer_flag = False
        geo_flag = False
        for row in tbody_general_info[0].find_all('tr'):
            tds_general_info = row.find_all('td')
            if len(tds_general_info) == 1:
                if tds_general_info[0].text == 'Поступление заказов от вебмастеров':
                    web_flag = True
                    offer_flag = geo_flag = False
                elif tds_general_info[0].text == 'Распределение заказов по офферам':
                    offer_flag = True
                    web_flag = geo_flag = False
                elif tds_general_info[0].text == 'Распределение заказов по странам':
                    geo_flag = True
                    web_flag = offer_flag = False
            else:
                if tds_general_info[0].text == '':
                    continue
                if web_flag:
                    self.webs_dict[tds_general_info[0].text] = tds_general_info[1].text
                elif offer_flag:
                    pass
                elif geo_flag:
                    self.geos_dict[tds_general_info[0].text] = tds_general_info[1].text
                    self.country_rows[tds_general_info[0].text] = []

    def compilation_xlsx_file(self):
        xlsx_file = openpyxl.load_workbook(template_address)
        report_sheet = xlsx_file['weekly report']
        country_sheet = xlsx_file['country Artem']
        report_sheet.title = f'WEEK {self.start_date} - {self.end_date}'
        prod_border = Border(left=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'),
                             right=Side(style='thin'))
        for web in self.webs_dict.keys():
            web_post = False
            web_row = report_sheet.max_row + 2
            web_rows_count = 0
            geo_row = web_row
            geo_rows_count = 0
            product_row = geo_row
            for geo in self.geos_dict.keys():
                print(web, geo)
                geo_post = False
                product_by_geo_response = self.session.get(
                    f'{os.getenv("platform_calls_analytics_link")}?from={self.start_date}&to={self.end_date}&wm={web}&geo={geo}')
                products_by_geo_soup = BeautifulSoup(product_by_geo_response.text, 'lxml')
                table = products_by_geo_soup.find(class_='table no-margin table-condensed table-bordered')
                tbody = table.find('tbody')
                rows = tbody.find_all('tr')
                for index in range(4, len(rows)):
                    tds = rows[index].find_all('td')
                    product_code = tds[0].text
                    if product_code == 'Распределение заказов по офферам':
                        continue
                    product_name = tds[1].text
                    success_leads = tds[9].text
                    web_rate = f'={tds[15].text}/{success_leads}'
                    skylead_rate = f'={tds[14].text}/{success_leads}'
                    if int(success_leads) > 0:
                        report_sheet.cell(row=product_row, column=column_index_from_string('C')).value = product_name
                        report_sheet.cell(row=product_row, column=column_index_from_string('D')).value = web_rate
                        report_sheet.cell(row=product_row, column=column_index_from_string('E')).value = skylead_rate
                        report_sheet.cell(row=product_row, column=3).border = prod_border
                        report_sheet.cell(row=product_row, column=4).border = prod_border
                        report_sheet.cell(row=product_row, column=5).border = prod_border
                        report_sheet.cell(row=product_row, column=column_index_from_string('I')).value = int(
                            success_leads)
                        report_sheet.cell(row=product_row,
                                          column=column_index_from_string(
                                              'G')).value = f'=SUM(E{product_row}-D{product_row})'
                        report_sheet.cell(row=product_row,
                                          column=column_index_from_string(
                                              'K')).value = f'=SUM(I{product_row}*E{product_row})'
                        report_sheet.cell(row=product_row,
                                          column=column_index_from_string(
                                              'L')).value = f'=SUM(I{product_row}*D{product_row})'
                        report_sheet.cell(row=product_row,
                                          column=column_index_from_string(
                                              'M')).value = f'=SUM(K{product_row}-L{product_row})'
                        product_row += 1
                        geo_rows_count += 1
                        geo_post = True
                        web_post = True

                if geo_post:
                    self.country_rows[geo].append((geo_row, geo_row + geo_rows_count - 1))
                    report_sheet.cell(row=geo_row, column=2).value = self.geos_dict.get(geo)
                    web_rows_count += geo_rows_count
                    geo_row += geo_rows_count
                geo_rows_count = 0
            if web_post:
                web_name = self.webs_dict[web]
                if self.webs_dict[web] == '[email\xa0protected]':
                    web_name = self.web_codes.get(web)
                    if not web_name:
                        web_name = f'web code: {web}'
                report_sheet.cell(row=web_row, column=1).value = web_name
                report_sheet.cell(row=web_row, column=column_index_from_string(
                    'O')).value = f'=SUM(M{web_row}:M{web_row + web_rows_count - 1})'

        """Filling in country sheet"""
        country_sheet.cell(row=2, column=1).value = f'{self.start_date} - {self.end_date}'
        it = 'SUM('
        hu = 'SUM('
        mx = 'SUM('
        ro = 'SUM('
        pl = 'SUM('
        de = 'SUM('
        fr = 'SUM('
        pe = 'SUM('

        if self.country_rows.get('it'):
            for rows_set in self.country_rows['it']:
                it += f"'WEEK {self.start_date} - {self.end_date}'!K{rows_set[0]}:K{rows_set[1]};"
            it += ')'
            country_sheet.cell(row=2, column=2).value = it
        else:
            country_sheet.cell(row=2, column=2).value = 0

        if self.country_rows.get('hu'):
            for rows_set in self.country_rows['hu']:
                hu += f"'WEEK {self.start_date} - {self.end_date}'!K{rows_set[0]}:K{rows_set[1]};"
            hu += ')'
            country_sheet.cell(row=2, column=3).value = hu
        else:
            country_sheet.cell(row=2, column=3).value = 0

        if self.country_rows.get('mx'):
            for rows_set in self.country_rows['mx']:
                mx += f"'WEEK {self.start_date} - {self.end_date}'!K{rows_set[0]}:K{rows_set[1]};"
            mx += ')'
            country_sheet.cell(row=2, column=4).value = mx
        else:
            country_sheet.cell(row=2, column=4).value = 0

        if self.country_rows.get('ro'):
            for rows_set in self.country_rows['ro']:
                ro += f"'WEEK {self.start_date} - {self.end_date}'!K{rows_set[0]}:K{rows_set[1]};"
            ro += ')'
            country_sheet.cell(row=2, column=5).value = ro
        else:
            country_sheet.cell(row=2, column=5).value = 0

        if self.country_rows.get('pl'):
            for rows_set in self.country_rows['pl']:
                pl += f"'WEEK {self.start_date} - {self.end_date}'!K{rows_set[0]}:K{rows_set[1]};"
            pl += ')'
            country_sheet.cell(row=2, column=6).value = pl
        else:
            country_sheet.cell(row=2, column=6).value = 0

        if self.country_rows.get('de'):
            for rows_set in self.country_rows['de']:
                de += f"'WEEK {self.start_date} - {self.end_date}'!K{rows_set[0]}:K{rows_set[1]};"
            de += ')'
            country_sheet.cell(row=2, column=7).value = de
        else:
            country_sheet.cell(row=2, column=7).value = 0

        if self.country_rows.get('fr'):
            for rows_set in self.country_rows['fr']:
                fr += f"'WEEK {self.start_date} - {self.end_date}'!K{rows_set[0]}:K{rows_set[1]};"
            fr += ')'
            country_sheet.cell(row=2, column=8).value = fr
        else:
            country_sheet.cell(row=2, column=8).value = 0

        if self.country_rows.get('pe'):
            for rows_set in self.country_rows['pe']:
                pe += f"'WEEK {self.start_date} - {self.end_date}'!K{rows_set[0]}:K{rows_set[1]};"
            pe += ')'
            country_sheet.cell(row=2, column=9).value = pe
        else:
            country_sheet.cell(row=2, column=9).value = 0

        xlsx_file.save(save_file_address)
        xlsx_file.close()


today = datetime.datetime.weekday(datetime.datetime.today())
start_date = datetime.datetime.today() - datetime.timedelta(weeks=1, days=today)
end_date = datetime.datetime.today() - datetime.timedelta(days=today + 1)

print(start_date.date(), end_date.date())

a = GeoDictGenerator()
a.compilation_geos_and_webs_lists(start_date.date(), end_date.date())
pprint(a.webs_dict)
pprint(a.geos_dict)
a.compilation_xlsx_file()
