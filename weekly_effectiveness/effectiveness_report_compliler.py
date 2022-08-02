import re
import requests
import datetime
from bs4 import BeautifulSoup
import os
import dotenv
import openpyxl
import json
from openpyxl.utils import column_index_from_string
from constants import *
from pathlib import Path

env_path = Path('.', '.env')
dotenv.load_dotenv(dotenv_path=env_path)


class EffectivenessReportCompiler:
    def __init__(self,
                 start_date: datetime.datetime,
                 end_date: datetime.datetime,
                 general_directory_path: str = '',
                 ):
        self.platform_info_dict = {}  # keeps info on the product name, 1)аналитика заказов -  кол, приняты, общий чек
        # , аналитика доставки - статус приняты, vat rate
        self.crm_geo_list_base = []
        self.crm_info_dict_base = {}
        self.crm_geo_list_mp = []
        self.crm_info_dict_mp = {}
        self.crm_geo_list_traff = []
        self.crm_info_dict_traff = {}
        self.start_date = start_date.date()
        self.end_date = end_date.date()
        self.general_directory_path = general_directory_path
        self.session = requests.Session()
        self.xlsx_file: openpyxl.Workbook = None  # crm export in xlsx format
        self.report_sheet: openpyxl.Workbook.active = None  # report sheet
        self.traffic_start_row = 1
        self.traffic_end_row = 24
        self.base_start_row = 27
        self.base_end_row = 59
        self.mp_start_row = 62
        self.mp_end_row = 94
        self.country_name: str = None  # Needed for the data change between the methods filling traff and base tables
        self.total_sells = {}
        self.data_for_eu_total_compiler = {}  # Format: 'geo':str : [traffic_total_row_num: int, base_total_row_num: int, mp_total_row_num: int]

    def get_list_of_geo_from_bl_export_files(self):
        for export_file in os.listdir(os.path.join(self.general_directory_path, 'exports_all')):
            if export_file.endswith('.xlsx') and not export_file.startswith('~$'):
                geo, chanel = export_file.split('_')
                chanel = chanel.split('.')[0]
                if chanel == 'base':
                    self.crm_geo_list_base.append(geo)
                elif chanel == 'mp':
                    self.crm_geo_list_mp.append(geo)
                elif chanel == 'traff':
                    self.crm_geo_list_traff.append(geo)

    def get_info_from_platform(self):
        """Getting info from Platform: аналитика заказов and followed by аналитика доставки."""
        self.session.post(os.getenv('platform_login_link'),
                          data={'in_user': os.getenv('platform_username'), 'in_pass': os.getenv('platform_password')})
        for geo in RATES.keys():  # iterating through each active country. RATES dict should have all of them as keys

            self.total_sells.setdefault(geo, 0)
            self.platform_info_dict.setdefault(geo, {})
            self.platform_info_dict[geo].setdefault('total sell',
                                                    0)  # creating nested dict for each geo, where the info on each product is stored, and general nested dict for the total sell info
            order_analytics_info = self.session.get(
                f'{os.getenv("platform_analytics_link")}?from={self.start_date}&to={self.end_date}&o=&c=&geo={geo}')
            order_analytics_soup = BeautifulSoup(order_analytics_info.text, 'lxml')
            order_analytics_table = order_analytics_soup.find('table')
            order_analytics_tbody = order_analytics_table.find('tbody')
            products_info_start = False  # flag indicating if the info on the products section of the  table is started
            products_info_end = False  # flag indicating if the info on the products section of the  table is ended
            product_name = None
            for tr in order_analytics_tbody.find_all('tr'):  # iterating through the rows
                if products_info_start and not products_info_end:  # if the products section is started and not ended yet, iterating through tds and adding info to the dict
                    if tr.text == 'Распределение заказов по странам':  # check if the row is the section end flag
                        products_info_end = True
                        continue
                    tds = tr.find_all(
                        'td')  # if the row is part of the products info section, get data and store to the dict
                    product_name = tds[1].text.split(' [')[0].lower()
                    leads = tds[2].text
                    sell = tds[9].text
                    vat_rate = VAT_RATES_DICT_FROM_BL.get(geo, {}).get(product_name)
                    self.platform_info_dict[geo].setdefault(product_name, {})
                    self.platform_info_dict[geo][product_name]['leads'] = int(leads)
                    self.platform_info_dict[geo][product_name]['sell'] = int(sell)
                    self.platform_info_dict[geo]['total sell'] += int(sell)
                    self.total_sells[geo] += int(sell)
                    if vat_rate:
                        self.platform_info_dict[geo][product_name]['vat_rate'] = vat_rate
                    else:
                        self.platform_info_dict[geo][product_name]['vat_rate'] = 'NO VAT INFO!'
                elif not products_info_start:  # check for the start flag
                    if tr.text == 'Распределение заказов по офферам':
                        products_info_start = True
                elif products_info_start and products_info_end:  # if the  section already started and ended, continue to the next geo. This line is probably never going into action as the continue is set for such case earlier.
                    continue

            # getting info from slylead: аналитика доставки
            delivery_analytics_info = self.session.get(f'{os.getenv("platform_delivery_info_link")}?'
                                                       f'from={self.start_date}&to={self.end_date}&geo={geo}')
            delivery_analytics_soup = BeautifulSoup(delivery_analytics_info.text, 'lxml')
            delivery_analytics_table = delivery_analytics_soup.find('table')
            delivery_analytics_tbody = delivery_analytics_table.find('tbody')
            products_info_start = False  # flag indicating if the info on the products section of the  table is started
            products_info_end = False  # flag indicating if the info on the products section of the  table is ended
            products_zero_bought = list(self.platform_info_dict[geo].keys())
            for tr in delivery_analytics_tbody.find_all('tr'):  # iterating through the rows
                if products_info_start and not products_info_end:  # if the products section is started and not ended yet, iterating through tds and adding info to the dict
                    if tr.text == 'Распределение заказов по странам':  # check if the row is the section end flag
                        products_info_end = True
                        continue
                    if self.platform_info_dict.get(geo, {}).get(product_name):
                        tds = tr.find_all(
                            'td')  # if the row is part of the products info section, get data and store to the dict
                        product_name = tds[1].text
                        bought = tds[7].text
                        income_info = tds[13].text
                        income_pattern = re.compile(r'\d+')
                        income_found = re.findall(income_pattern, income_info)
                        if len(income_found) > 0:
                            income = income_found[0]
                        else:
                            income = 0
                elif not products_info_start:  # check for the start flag
                    if tr.text == 'Распределение заказов по офферам':
                        products_info_start = True
                elif products_info_start and products_info_end:  # if the  section already started and ended, continue to the next geo. This line is probably never going into action as the continue is set for such case earlier.
                    continue

            # for those products not in аналитика заказов adding bought: 0
            # print(products_zero_bought)
            for product_zero_bought in products_zero_bought:
                if not product_zero_bought == 'total sell':
                    self.platform_info_dict[geo][product_zero_bought]['bought'] = 0

            # get adv expenditures per product from аналитика звонков
            calls_data = self.session.get(
                f'{os.getenv("platform_calls_analytics_link")}?from={self.start_date}&to={self.end_date}&o=&c=&geo={geo}')
            calls_data_soup = BeautifulSoup(calls_data.text, 'lxml')
            calls_data_table = calls_data_soup.find('table')
            calls_data_tbody = calls_data_table.find('tbody')
            geo_products = list(self.platform_info_dict[geo].keys())
            for tr in calls_data_tbody.find_all('tr'):  # iterating through the rows
                tds = list(tr.find_all('td'))
                if len(tds) > 1:
                    if tds[1].text in geo_products:
                        self.platform_info_dict[geo][tds[1].text]['adv'] = int(tds[-3].text)


    def get_info_from_bl_export(self, file_path: str, chanel: str, geo: str):
        """The method works with the export from the CRM (login=base) in the xlsx format. The method will be called inside the
         other method that will iterate through the xlsx files in the folder, but all the logic of work with xlsx is
         here."""

        base_xlsx_file = openpyxl.load_workbook(
            os.path.join(self.general_directory_path, f'exports_all', file_path))
        info_dict = {}  # empties the dict, as the data is already added to the final report file
        info_dict.setdefault(geo, {})  # creating nested geo dict
        sheet = base_xlsx_file.active

        # Removing adv cost for the secondary products for orders in traff file
        if chanel == 'traff':
            prev_order_id_adv = 0
            for order_row_adv in range(2, sheet.max_row + 1):
                order_id_adv = sheet.cell(row=order_row_adv, column=1).value
                if order_id_adv == prev_order_id_adv:
                    sheet.cell(row=order_row_adv, column=column_index_from_string('O')).value = 0
                prev_order_id_adv = order_id_adv

            # Saving changes and reopening file
            base_xlsx_file.save(os.path.join(self.general_directory_path, f'exports_all', file_path))
            base_xlsx_file = openpyxl.load_workbook(
                os.path.join(self.general_directory_path, f'exports_all', file_path))
            sheet = base_xlsx_file.active

        previous_order_id = 0
        primary_product = ''
        for row in range(2, sheet.max_row + 1):  # iterating through rows, adding info to the dict earlier
            product = sheet.cell(row=row, column=5).value.strip().title()
            self.total_sells.setdefault(geo, 0)
            info_dict.setdefault(geo, {})
            info_dict[geo].setdefault(product, {'sell, pcs': 0, 'sell, eur': 0, 'bought, pcs': 0,
                                                'income, brutto': 0, 'vat, eur': 0, 'vat rate': 0,
                                                'products count, pcs': 0, 'adv cost': 0,
                                                'bought orders with product': 0, 'courier cost': 0,
                                                'income delivery, eur': 0, 'vat delivery, eur': 0})
            info_dict.setdefault('total sell', 0)
            product_quantity = sheet.cell(row=row, column=column_index_from_string('H')).value
            price_brutto = round(
                float(sheet.cell(row=row, column=column_index_from_string('I')).value) / RATES.get(geo) *
                product_quantity, 2)
            vat_amount = round(float(sheet.cell(row=row, column=column_index_from_string('M')).value) / RATES.get(geo) *
                               product_quantity, 2)
            order_id = sheet.cell(row=row, column=1).value
            adv_cost = sheet.cell(row=row, column=column_index_from_string('O')).value
            status = sheet.cell(row=row, column=column_index_from_string('N')).value
            delivery_brutto = round(float(sheet.cell(row=row, column=column_index_from_string('K')).value) /
                                    RATES.get(geo), 2)

            if order_id != previous_order_id:  # checking if new row data is not referred to the the order id from the previous line. If order id is new, 1 'sell, pcs' is added
                info_dict['total sell'] += 1
                self.total_sells[geo] += 1
                info_dict[geo][product]['sell, pcs'] += 1
                primary_product = product
                if info_dict[geo][product]['vat rate'] == 0:
                    single_product_vat_amount = float(sheet.cell(row=row, column=column_index_from_string('M')).value)
                    single_product_brutto_price = float(sheet.cell(row=row, column=column_index_from_string('I')).value)
                    if single_product_brutto_price == 0:
                        info_dict[geo][product]['vat rate'] = 0
                    else:
                        info_dict[geo][product]['vat rate'] = single_product_vat_amount / \
                                                              (single_product_brutto_price - single_product_vat_amount) \
                                                              * 100
                info_dict[geo][primary_product]['sell, eur'] += delivery_brutto

            info_dict[geo][primary_product]['sell, eur'] += price_brutto

            if adv_cost:
                info_dict[geo][product]['adv cost'] += adv_cost

            if status.lower() in ['delivered', 'money received']:
                if order_id != previous_order_id:
                    info_dict[geo][product]['bought, pcs'] += 1
                    info_dict[geo][product]['income, brutto'] += price_brutto
                    info_dict[geo][product]['vat, eur'] += vat_amount
                    info_dict[geo][product]['products count, pcs'] += product_quantity
                    info_dict[geo][product]['bought orders with product'] += 1

                    delivery_vat = delivery_brutto * DELIVERY_VAT_RATES[geo] / (100 + DELIVERY_VAT_RATES[geo])
                    delivery_net = delivery_brutto - delivery_vat
                    info_dict[geo][product]['income delivery, eur'] += delivery_brutto
                    info_dict[geo][product]['vat delivery, eur'] += delivery_vat

                else:
                    info_dict[geo][primary_product]['income, brutto'] += price_brutto
                    info_dict[geo][primary_product]['vat, eur'] += vat_amount
                    info_dict[geo][primary_product]['products count, pcs'] += product_quantity
            previous_order_id = order_id

        if chanel == 'base':
            self.crm_info_dict_base = info_dict

        elif chanel == 'mp':
            self.crm_info_dict_mp = info_dict

        elif chanel == 'traff':
            self.crm_info_dict_traff = info_dict

    def fill_in_crm_table(self, geo: str, chanel: str):
        """Method takes data from the CRM data dict and fills in the base and mp table"""
        crm_table_start = 0
        crm_table_end = 0
        info_dict = None
        if chanel == 'base':
            crm_table_start = self.base_start_row
            crm_table_end = self.base_end_row
            info_dict = self.crm_info_dict_base
        elif chanel == 'mp':
            crm_table_start = self.mp_start_row
            crm_table_end = self.mp_end_row
            info_dict = self.crm_info_dict_mp
        elif chanel == 'traff':
            crm_table_start = self.traffic_start_row
            crm_table_end = self.traffic_end_row
            info_dict = self.crm_info_dict_traff

        # filling in headers of the tables
        country_name = COUNTRIES.get(geo)
        print(geo, chanel)
        print('INFO DICT!!!!!!', info_dict)
        self.report_sheet.cell(row=crm_table_start, column=1).value = country_name
        self.report_sheet.cell(row=crm_table_start, column=2).value = f'{self.start_date} - {self.end_date}'

        # filling in data to the base column from the self.self.crm_info_dict

        crm_product_row = crm_table_start + 2
        for product in info_dict[geo]:
            if not product == 'total sell':
                if info_dict[geo][product]['sell, pcs'] > 0:
                    self.report_sheet.cell(row=crm_product_row,
                                           column=column_index_from_string('A')).value = product
                    if chanel == 'traff':
                        leads = self.platform_info_dict[geo].get(product.lower(), {}).get('leads')
                        if leads:
                            self.report_sheet.cell(row=crm_product_row,
                                                   column=column_index_from_string('B')).value = leads
                        else:
                            self.report_sheet.cell(row=crm_product_row,
                                                   column=column_index_from_string('B')).value = 0
                        self.report_sheet.cell(row=crm_product_row,
                                               column=column_index_from_string(
                                                   'E')).value = f'=C{crm_product_row}/' \
                                                                 f'B{crm_product_row}'

                    self.report_sheet.cell(row=crm_product_row, column=column_index_from_string('C')).value = \
                        int(info_dict[geo][product]['sell, pcs'])
                    self.report_sheet.cell(row=crm_product_row, column=column_index_from_string('D')).value = \
                        float(info_dict[geo][product]['sell, eur'])
                    if info_dict[geo][product]['bought, pcs'] == 0:
                        for column_letter in ['F', 'G', 'H', 'I', 'L', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']:
                            self.report_sheet.cell(row=crm_product_row,
                                                   column=column_index_from_string(column_letter)).value = 0
                        self.report_sheet.cell(row=crm_product_row,
                                               column=column_index_from_string('P')).value = -1
                    else:
                        self.report_sheet.cell(row=crm_product_row, column=column_index_from_string('F')).value = \
                            int(info_dict[geo][product]['bought, pcs'])
                        self.report_sheet.cell(row=crm_product_row, column=column_index_from_string('G')).value = \
                            f'={info_dict[geo][product]["income, brutto"]} + {info_dict[geo][product]["income delivery, eur"]}'
                        self.report_sheet.cell(row=crm_product_row, column=column_index_from_string('H')).value = \
                            f'=F{crm_product_row}/C{crm_product_row}'

                        if geo in REDUCED_VAT_RATE:
                            self.report_sheet.cell(row=crm_product_row,
                                                   column=column_index_from_string('I')).value = \
                                f'=({float(info_dict[geo][product]["vat, eur"])} + {info_dict[geo][product]["vat delivery, eur"]})*0.2'
                        else:
                            self.report_sheet.cell(row=crm_product_row, column=column_index_from_string('I')).value = \
                                f'={float(info_dict[geo][product]["vat, eur"])} + {info_dict[geo][product]["vat delivery, eur"]}'
                        if chanel == 'traff':
                            self.report_sheet.cell(row=crm_product_row,
                                                   column=column_index_from_string('K')).value = \
                                f'={float(info_dict[geo][product]["adv cost"])}/{RATES["eur/usd"]}'
                        self.report_sheet.cell(row=crm_product_row, column=column_index_from_string('L')).value = \
                            f'={info_dict[geo][product]["products count, pcs"]}*{OPEX_GOODS_COST[geo]["goods cost"]}'
                        self.report_sheet.cell(row=crm_product_row, column=column_index_from_string('P')).value = \
                            f'=O{crm_product_row}/D{crm_product_row}'
                        self.report_sheet.cell(row=crm_product_row, column=column_index_from_string('Q')).value = \
                            f'=D{crm_product_row}/C{crm_product_row}'
                        # if info_dict[geo][product]['bought, pcs'] > 0:
                        # base_product_vat_rate = info_dict[geo][product]["vat, eur"] / (
                        #         info_dict[geo][product]["income, brutto"] -
                        #         info_dict[geo][product]["vat, eur"]) * 100
                        # else:
                        #     base_product_vat_rate = 'NO VAT INFO!!!'
                        self.report_sheet.cell(row=crm_product_row, column=column_index_from_string('R')).value = \
                            f'=-(J{crm_product_row}+K{crm_product_row}+L{crm_product_row}+M{crm_product_row})/(0-H{crm_product_row}*(1-{info_dict[geo][product]["vat rate"]}/(100+{info_dict[geo][product]["vat rate"]})))/C{crm_product_row}'
                        self.report_sheet.cell(row=crm_product_row, column=column_index_from_string('S')).value = \
                            f'=-(J{crm_product_row}+K{crm_product_row}+L{crm_product_row}+M{crm_product_row})/(0.1-H{crm_product_row}*(1-{info_dict[geo][product]["vat rate"]}/(100+{info_dict[geo][product]["vat rate"]})))/C{crm_product_row}'
                        # if info_dict[geo][product]['bought, pcs'] > 0:
                        y_column_value = info_dict[geo][product]['products count, pcs'] / \
                                         info_dict[geo][product]['bought orders with product']
                        # else:
                        #     y_column_value = 0
                        self.report_sheet.cell(row=crm_product_row, column=column_index_from_string('T')).value = \
                            f'=(0*D{crm_product_row}+C{crm_product_row}*{COURIER_TARIFFS[geo]["sent"]}+C{crm_product_row}*{COURIER_TARIFFS[geo]["return"]}+K{crm_product_row}+M{crm_product_row})/(D{crm_product_row}-D{crm_product_row}*{info_dict[geo][product]["vat rate"]}/(100+{info_dict[geo][product]["vat rate"]})-C{crm_product_row}*{y_column_value}*{OPEX_GOODS_COST[geo]["goods cost"]}+C{crm_product_row}*{COURIER_TARIFFS[geo]["return"]})'
                        self.report_sheet.cell(row=crm_product_row, column=column_index_from_string('U')).value = \
                            f'=(0.1*D{crm_product_row}+C{crm_product_row}*{COURIER_TARIFFS[geo]["sent"]}+C{crm_product_row}*{COURIER_TARIFFS[geo]["return"]}+K{crm_product_row}+M{crm_product_row})/(D{crm_product_row}-D{crm_product_row}*{info_dict[geo][product]["vat rate"]}/(100+{info_dict[geo][product]["vat rate"]})-C{crm_product_row}*{y_column_value}*{OPEX_GOODS_COST[geo]["goods cost"]}+C{crm_product_row}*{COURIER_TARIFFS[geo]["return"]})'
                        self.report_sheet.cell(row=crm_product_row, column=column_index_from_string('V')).value = \
                            f'=(K{crm_product_row}+M{crm_product_row})/(Q{crm_product_row}*H{crm_product_row}-Q{crm_product_row}*H{crm_product_row}*{info_dict[geo][product]["vat rate"]}/(100+{info_dict[geo][product]["vat rate"]})-({COURIER_TARIFFS[geo]["sent"]}+{COURIER_TARIFFS[geo]["return"]}*(1-H{crm_product_row}))-{y_column_value}*{OPEX_GOODS_COST[geo]["goods cost"]}*H{crm_product_row}-0*Q{crm_product_row})'
                        self.report_sheet.cell(row=crm_product_row, column=column_index_from_string('W')).value = \
                            f'=(K{crm_product_row}+M{crm_product_row})/(Q{crm_product_row}*H{crm_product_row}-Q{crm_product_row}*H{crm_product_row}*{info_dict[geo][product]["vat rate"]}/(100+{info_dict[geo][product]["vat rate"]})-({COURIER_TARIFFS[geo]["sent"]}+{COURIER_TARIFFS[geo]["return"]}*(1-H{crm_product_row}))-{y_column_value}*{OPEX_GOODS_COST[geo]["goods cost"]}*H{crm_product_row}-0.1*Q{crm_product_row})'

                    self.report_sheet.cell(row=crm_product_row, column=column_index_from_string('J')).value = \
                        f'=C{crm_product_row}*{COURIER_TARIFFS[geo]["sent"]}+C{crm_product_row}*(1-H{crm_product_row})*{COURIER_TARIFFS[geo]["return"]}'
                    self.report_sheet.cell(row=crm_product_row, column=column_index_from_string('M')).value = \
                        f'=C{crm_product_row}*{OPEX_GOODS_COST[geo]["opex"]}'
                    self.report_sheet.cell(row=crm_product_row, column=column_index_from_string('N')).value = \
                        f'=I{crm_product_row}+J{crm_product_row}+K{crm_product_row}+L{crm_product_row}+M{crm_product_row}'
                    self.report_sheet.cell(row=crm_product_row, column=column_index_from_string('O')).value = \
                        f'=G{crm_product_row}-N{crm_product_row}'

                    crm_product_row += 1

        # Deleting all empty rows left between the last product row and total row
        self.report_sheet.delete_rows(crm_product_row, crm_table_end - crm_product_row)
        crm_table_end = crm_product_row

        if chanel == 'traff':
            self.base_start_row = crm_table_end + 3
            self.base_end_row = self.base_start_row + 32
            self.mp_start_row = self.base_end_row + 3
            self.mp_end_row = self.mp_start_row + 32
            self.data_for_eu_total_compiler.setdefault(geo, {})
            self.data_for_eu_total_compiler[geo].setdefault('traff', crm_table_end)

        elif chanel == 'base':
            # resetting values for mp table range since the rows above were deleted.
            self.mp_start_row = crm_table_end + 3
            self.mp_end_row = self.mp_start_row + 32
            self.data_for_eu_total_compiler.setdefault(geo, {})
            self.data_for_eu_total_compiler[geo].setdefault('base', crm_table_end)

        elif chanel == 'mp':
            self.data_for_eu_total_compiler.setdefault(geo, {})
            self.data_for_eu_total_compiler[geo].setdefault('mp', crm_table_end)

        # Filling total row for the base table
        # base_start_row + 3 = the first row with the product info in the base table
        # base_end_row - 1 = the last row with the product info in the base table

        # filling in cells with the value for total row as the sum of the values in the column
        for column_letter in ['C', 'D', 'F', 'G', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'V', 'W']:
            self.report_sheet.cell(row=crm_table_end, column=column_index_from_string(column_letter)).value = \
                f'=SUM({column_letter}{crm_table_start + 2}:{column_letter}{crm_table_end - 1})'

        if chanel == 'traff':
            self.report_sheet.cell(row=crm_table_end, column=column_index_from_string('B')).value = \
                f'=SUM({"B"}{crm_table_start + 2}:{"B"}{crm_table_end - 1})'
            self.report_sheet.cell(row=crm_table_end, column=column_index_from_string('E')).value = \
                f'=C{crm_table_end}/B{crm_table_end}'

        self.report_sheet.cell(row=crm_table_end, column=column_index_from_string('H')).value = \
            f'=F{crm_table_end}/C{crm_table_end}'
        self.report_sheet.cell(row=crm_table_end, column=column_index_from_string('P')).value = \
            f'=O{crm_table_end}/D{crm_table_end}'
        self.report_sheet.cell(row=crm_table_end, column=column_index_from_string('Q')).value = \
            f'=D{crm_table_end}/C{crm_table_end}'
        self.report_sheet.cell(row=crm_table_end, column=column_index_from_string('R')).value = \
            f'SUMPRODUCT(F{crm_table_start + 2}:F{crm_table_end - 1};R{crm_table_start + 2}:R{crm_table_end - 1})/F{crm_table_end}'
        self.report_sheet.cell(row=crm_table_end, column=column_index_from_string('S')).value = \
            f'SUMPRODUCT($F${crm_table_start + 2}:$F${crm_table_end - 1};S${crm_table_start + 2}:S${crm_table_end - 1})/$F${crm_table_end}'
        self.report_sheet.cell(row=crm_table_end, column=column_index_from_string('T')).value = \
            f'SUMPRODUCT($D${crm_table_start + 2}:$D${crm_table_end - 1};T${crm_table_start + 2}:T${crm_table_end - 1})/$D${crm_table_end}'
        self.report_sheet.cell(row=crm_table_end, column=column_index_from_string('U')).value = \
            f'SUMPRODUCT($D${crm_table_start + 2}:$D${crm_table_end - 1};U${crm_table_start + 2}:U${crm_table_end - 1})/$D${crm_table_end}'

    def act(self):
        """Method consicts of the logic opf calling all other methopds of the class in order to make all the
        reports for all geos."""

        self.get_list_of_geo_from_bl_export_files()
        self.get_info_from_platform()

        # running whole logic for each geo
        self.xlsx_file = openpyxl.load_workbook(
            os.path.join('/Users/artem/Desktop/работа/NC/NEW/effectiveness_report/template_combined.xlsx'))
        for geo in list(COUNTRIES.keys()):
            self.report_sheet = self.xlsx_file[f'{geo}']
            if geo in self.crm_geo_list_traff:
                self.get_info_from_bl_export(f'{geo}_traff.xlsx', chanel='traff', geo=geo)
            if geo in self.crm_geo_list_base:
                self.get_info_from_bl_export(f'{geo}_base.xlsx', chanel='base', geo=geo)
            if geo in self.crm_geo_list_mp:
                self.get_info_from_bl_export(f'{geo}_mp.xlsx', chanel='mp', geo=geo)

            # self.fill_in_traff_table(geo)
            if geo in self.crm_geo_list_traff:
                self.fill_in_crm_table(geo, chanel='traff')
            if geo in self.crm_geo_list_base:
                self.fill_in_crm_table(geo, chanel='base')
            if geo in self.crm_geo_list_mp:
                self.fill_in_crm_table(geo, chanel='mp')

            os.makedirs(os.path.join(self.general_directory_path, 'reports',
                                     f'{self.start_date}-{self.end_date}',
                                     'script_report', ), exist_ok=True)

            # print(geo, 'OPEX:', OPEX_ADV_GOODS_COST[geo]['opex'] * self.total_sells[geo])
            # back to custom data
            self.crm_info_dict_base = {}
            self.crm_info_dict_mp = {}
            self.traffic_start_row = 1
            self.traffic_end_row = 24
            self.base_start_row = 27
            self.base_end_row = 59
            self.mp_start_row = 62
            self.mp_end_row = 94

        path_to_file_with_formulas = os.path.join(self.general_directory_path, 'reports',
                                                  f'{self.start_date}-{self.end_date}',
                                                  'script_report',
                                                  f'eff_{self.start_date}_{self.end_date}_formulas.xlsx')

        self.xlsx_file.save(path_to_file_with_formulas)
        self.xlsx_file.close()

        self.data_for_eu_total_compiler['start_date'] = str(self.start_date)
        self.data_for_eu_total_compiler['end_date'] = str(self.end_date)
        self.data_for_eu_total_compiler['formulas_file_path'] = path_to_file_with_formulas
        totals_json_string = json.dumps(self.data_for_eu_total_compiler)

        with open('data_passed.json', 'w') as totals_json_file:
            totals_json_file.write(totals_json_string)
