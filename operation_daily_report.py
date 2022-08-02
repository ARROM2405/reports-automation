import calendar
import re
from pathlib import Path

import bs4
import requests
import openpyxl
from openpyxl.utils import column_index_from_string
import datetime
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os
import dotenv

env_path = Path('.', '.env')
dotenv.load_dotenv(dotenv_path=env_path)

file_address = input('Please input file address.\n')

RATES = {}  # Dict in the format {'geo_name'(str): currency rate to EUR(dec)}
CALLCENTER_CAPACITY = {}  # Callcenter capacity in format {'geo'(str): leads_number(int)}
TARGETS = {}  # Targets in format {'geo name'(str): {'product name'(str): (check target(int), approve rate(int))}}

MONTHS = {1: 'Январь',
          2: 'Февраль',
          3: 'Март',
          4: 'Апрель',
          5: 'Май',
          6: 'Июнь',
          7: 'Июль',
          8: 'Август',
          9: 'Сентябрь',
          10: 'Октябрь',
          11: 'Ноябрь',
          12: 'Декабрь'}


def delivery_rates(geo, start_date, end_date):
    session = requests.Session()
    if geo == 'mx':
        session.get(os.getenv('mx_recover_link'))
        geo_info = session.get(f'{os.getenv("mx_delivery_info_link")}?from={start_date}&to={end_date}&geo=mx')
    else:
        session.post(os.getenv('platform_login_link'),
                     data={'in_user': os.getenv('platform_username'), 'in_pass': os.getenv('platform_password')})
        geo_info = session.get(
            f'{os.getenv("platform_delivery_info_link")}?from={start_date}&to={end_date}&o=&c=&geo={geo}')
    soup = bs4.BeautifulSoup(geo_info.text, 'lxml')
    table = soup.find('tbody')
    tr = table.find('tr')
    tds = tr.find_all('td')
    try:
        return tds[9].text
    except:
        return 'No info'


class OffersDictGenerator:
    """Gets JSON with the list of offers (products)."""

    def __init__(self, url: str) -> None:
        self.offers_list = requests.get(os.getenv('platform_offers_api_link'))
        self.offers_json = self.offers_list.json()


class StatisticsProcessor:
    def __init__(self, start_date: str, end_date: str = None):
        self.start_date = start_date
        self.end_date = end_date
        if self.end_date == None:
            self.end_date = self.start_date
        self.geo_dict = dict()
        self.country_rows = list()
        self.total_row = 0
        self.opened_xlsx_file = None
        self.report_sheet = None
        self.mexico_average_check = 0

    def fill_geo_dict(self):
        """Gets the list of active GEO for the dates and puts them into a dict format: 'GEO code': 'GEO Name'."""
        geos_json = requests.get(
            f'{os.getenv("platform_active_geo_api_link")}&from={self.start_date}&to={self.end_date}')
        for k, v in geos_json.json().items():
            if k in RATES.keys():
                if v['name'] is not None:
                    geo_name = v['name']
                    geo_approve_rate = v['apps']
                    geo_approves_data = v['phase'].get('3')
                    geo_approves_sum = 0
                    geo_average_sum = 0
                    if geo_approves_data:
                        geo_approves_values = geo_approves_data.get('cash')
                        geo_total_value = list(geo_approves_data.get('cash').values())[0]
                        geo_average_sum = geo_total_value[1] / geo_total_value[0]
                        for v in geo_approves_values.values():
                            geo_approves_sum = v[1]
                    self.geo_dict[k] = [geo_name, geo_approve_rate, geo_approves_sum, geo_average_sum]
        print(self.geo_dict)
        return self.geo_dict

    def connect_with_xlsx_file_and_sheet_create(self):
        self.opened_xlsx_file = xls_file = openpyxl.load_workbook(file_address)
        self.report_sheet = sheet = self.opened_xlsx_file.create_sheet(title=f'{self.start_date}', index=0)
        sheet.cell(row=1, column=1).value = self.start_date

    def country_row_fill_in(self, country_row, geo, geo_name, products_num):
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        bold = Font(bold=True)
        print(geo)
        """Filling in country row with formulas and values."""
        self.report_sheet.cell(row=country_row, column=1).value = geo_name
        self.report_sheet.cell(row=country_row,
                               column=2).value = f'=SUM(B{country_row + 1}:B{country_row + products_num})'
        self.report_sheet.cell(row=country_row, column=3).value = f'=B{country_row}/D{country_row}'
        self.report_sheet.cell(row=country_row, column=4).value = CALLCENTER_CAPACITY[geo]
        self.report_sheet.cell(row=country_row,
                               column=5).value = f'=SUM(E{country_row + 1}:E{country_row + products_num})'
        self.report_sheet.cell(row=country_row, column=6).value = f'=E{country_row}/B{country_row}'
        self.report_sheet.cell(row=country_row, column=7).value = f'=F{country_row}-O{country_row}'
        self.report_sheet.cell(row=country_row, column=9).value = f'=H{country_row}-N{country_row}'
        self.report_sheet.cell(row=country_row, column=12).value = f'=K{country_row}-J{country_row}'

        """Filling in targets"""
        self.report_sheet.cell(row=country_row,
                               column=column_index_from_string(
                                   'N')).value = f'=AVERAGE(N{country_row + 1}:N{country_row + products_num})'
        self.report_sheet.cell(row=country_row,
                               column=column_index_from_string(
                                   'O')).value = f'=AVERAGE(O{country_row + 1}:O{country_row + products_num})'

        if self.geo_dict[geo][2] != 0:
            if geo != 'mx':
                self.report_sheet.cell(row=country_row,
                                       column=8).value = f'=({self.geo_dict[geo][2]}/E{country_row})/{RATES.get(geo)}'
            else:
                self.report_sheet.cell(row=country_row,
                                       column=8).value = self.mexico_average_check / RATES['mx']


        else:
            self.report_sheet.cell(row=country_row, column=8).value = 0

        """Setting bold font for country row."""

        for column in range(1, column_index_from_string('L') + 1):
            self.report_sheet.cell(row=country_row, column=column).font = bold

        for column_letter in ['N', 'O']:
            self.report_sheet.cell(row=country_row, column=column_index_from_string(column_letter)).font = bold
            self.report_sheet.cell(row=country_row, column=column_index_from_string(column_letter)).fill = yellow_fill
        self.report_sheet.cell(row=country_row,
                               column=column_index_from_string('N')).number_format = '0'
        self.report_sheet.cell(row=country_row,
                               column=column_index_from_string('O')).number_format = '0%'

    def product_row_fill_in(self, info_by_geo, geo, country_row, products_num):
        """Filling in product row with formulas and values."""
        for product in info_by_geo.json().values():
            product_name = product['name']
            product_row = self.report_sheet.max_row + 1
            self.report_sheet.cell(row=product_row, column=1).value = product_name
            self.report_sheet.cell(row=product_row, column=2).value = product['count']
            phase_3 = product['phase'].get('3')
            if phase_3:
                self.report_sheet.cell(row=product_row, column=5).value = phase_3.get('count')
                self.report_sheet.cell(row=product_row, column=6).value = f'=E{product_row}/B{product_row}'
                total_paid_info = list(phase_3.get('cash').values())[0]
                average = round(total_paid_info[1] / total_paid_info[0])
                self.report_sheet.cell(row=product_row,
                                       column=8).value = f'=({average})/{RATES.get(geo)}'
            else:
                self.report_sheet.cell(row=product_row, column=6).value = 0
                self.report_sheet.cell(row=product_row, column=5).value = 0
                self.report_sheet.cell(row=product_row, column=8).value = 0
            self.report_sheet.cell(row=product_row, column=7).value = f'=F{product_row}-O{product_row}'
            self.report_sheet.cell(row=product_row, column=9).value = f'=H{product_row}-N{product_row}'

            """Filling in targets."""
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            geo_targets = TARGETS.get(geo)
            if geo_targets:
                product_targets = geo_targets.get(product_name)
                if product_targets:
                    self.report_sheet.cell(row=product_row,
                                           column=column_index_from_string('N')).fill = yellow_fill
                    self.report_sheet.cell(row=product_row,
                                           column=column_index_from_string('N')).value = product_targets[0]
                    self.report_sheet.cell(row=product_row,
                                           column=column_index_from_string('O')).fill = yellow_fill
                    self.report_sheet.cell(row=product_row,
                                           column=column_index_from_string('O')).value = product_targets[1] / 100
                    self.report_sheet.cell(row=product_row,
                                           column=column_index_from_string('O')).number_format = '0%'

        self.report_sheet.row_dimensions.group(country_row + 1, country_row + products_num, hidden=True)

    def sheet_headers_filling_in_and_columns_dimensions(self):
        bold = Font(bold=True)
        self.report_sheet.cell(row=1, column=1).value = self.start_date

        """Columns width setup."""
        self.report_sheet.column_dimensions['A'].width = 25
        self.report_sheet.column_dimensions['B'].width = 9.84
        self.report_sheet.column_dimensions['C'].width = 9.84
        self.report_sheet.column_dimensions['D'].width = 11
        self.report_sheet.column_dimensions['E'].width = 9.84
        self.report_sheet.column_dimensions['F'].width = 9.84
        self.report_sheet.column_dimensions['G'].width = 10.83
        self.report_sheet.column_dimensions['H'].width = 12.67
        self.report_sheet.column_dimensions['I'].width = 12
        self.report_sheet.column_dimensions['J'].width = 9.83
        self.report_sheet.column_dimensions['K'].width = 9.83
        self.report_sheet.column_dimensions['L'].width = 9.83
        self.report_sheet.column_dimensions['N'].width = 10
        self.report_sheet.column_dimensions['O'].width = 14.33

        """Borders style."""
        row_1_border = Border(bottom=Side(style='thin'))
        row_2_border = Border(left=Side(style='thin'),
                              bottom=Side(style='thin'),
                              right=Side(style='thin'))

        """Columns range of the main table."""
        column_range = range(1, 13)

        """Color fill in."""
        grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        """Main header font."""
        bold_header = Font(size=14, bold=True)

        """Setting main table headers font."""
        for column in column_range:
            self.report_sheet.cell(row=1, column=column).fill = grey_fill
            self.report_sheet.cell(row=1, column=column).border = row_1_border
            self.report_sheet.cell(row=1, column=column).font = bold_header
            self.report_sheet.cell(row=2, column=column).fill = grey_fill
            self.report_sheet.cell(row=2, column=column).border = row_2_border
            self.report_sheet.cell(row=2, column=column).font = bold_header
            self.report_sheet.cell(row=2, column=column).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                               vertical='center')
        """Setting plan table headers font """
        for column in range(14, 16):
            self.report_sheet.cell(row=1, column=column).fill = grey_fill
            self.report_sheet.cell(row=1, column=column).font = bold_header
            self.report_sheet.cell(row=2, column=column).fill = yellow_fill
            self.report_sheet.cell(row=2, column=column).border = row_2_border
            self.report_sheet.cell(row=2, column=column).font = bold_header
            self.report_sheet.cell(row=2, column=column).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                               vertical='center')

        """Setting header row height."""
        self.report_sheet.row_dimensions[1].height = 22
        self.report_sheet.row_dimensions[2].height = 80

        """Filling in headers."""
        self.report_sheet.cell(row=2, column=1).value = 'Страна - Товар'
        self.report_sheet.cell(row=2, column=2).value = 'Лиды (шт).'
        self.report_sheet.cell(row=2, column=3).value = 'Загрузка КЦ (%)'
        self.report_sheet.cell(row=2, column=4).value = 'Мощность КЦ'
        self.report_sheet.cell(row=2, column=5).value = 'Апрув (шт)'
        self.report_sheet.cell(row=2, column=6).value = 'Апрув\n(%)'
        self.report_sheet.cell(row=2, column=7).value = 'Аппрув к\nmin\nуровню'
        self.report_sheet.cell(row=2, column=8).value = 'Ср. чек\n€'
        self.report_sheet.cell(row=2, column=9).value = 'Ср. чек\nк целевому\n+/-  €'
        self.report_sheet.cell(row=2, column=10).value = 'К отправке'
        self.report_sheet.cell(row=2, column=11).value = 'Отправлено(шт)'
        self.report_sheet.cell(row=2, column=12).value = 'Отправка\n+\- шт'
        self.report_sheet.cell(row=2, column=14).value = 'План\nср. чек'
        self.report_sheet.cell(row=2, column=15).value = 'Минимальный показатель апрува'
        self.report_sheet.cell(row=1, column=14).value = 'Целевые показатели'

    def compiling_with_mex_info(self, country_row):
        """Getting info for Mexico"""
        session = requests.Session()
        session.get(os.getenv('mx_recover_link'))
        mexico_analytics_info = session.get(
            f'{os.getenv("mx_analytics_info_link")}?from={self.start_date}&to={self.end_date}&geo=mx')
        soup = bs4.BeautifulSoup(mexico_analytics_info.text, 'lxml')
        """Getting table with the data and iterating through rows and columns"""
        table = soup.find('tbody')
        products_start = False
        products_end = False
        first_row = table.find('tr')
        country_average_check_value = first_row.find_all('td')[-2].text
        country_average_check_sum = int(re.findall(re.compile(r'\d+'), country_average_check_value)[0])
        mexico_prod_dict = dict()
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        total_mexico_approves = 0
        total_mexico_check = 0
        for row in table.find_all('tr')[1:]:
            tds = row.find_all('td')
            if len(tds) == 1:
                if tds[0].text == 'Распределение заказов по офферам':
                    products_start = True
                    continue
                elif tds[0].text == 'Распределение заказов по странам':
                    products_end = True
                    break
            if products_start and not products_end:
                product_name = tds[1].text
                leads_num = int(tds[2].text)
                approves_num = int(tds[9].text)
                product_average_check_sum = 0
                if approves_num > 0:
                    product_average_check_value = tds[13].text
                    product_average_check_sum = int(re.findall(re.compile(r'\d+'), product_average_check_value)[0])
                mexico_prod_dict[product_name] = [leads_num, approves_num, product_average_check_sum]

                for product in mexico_prod_dict.keys():
                    if mexico_prod_dict[product][1] != 0:
                        total_mexico_approves += mexico_prod_dict[product][1]
                        total_mexico_check += mexico_prod_dict[product][2] * mexico_prod_dict[product][1]

                self.mexico_average_check = total_mexico_check / total_mexico_approves

        self.country_row_fill_in(country_row=country_row, geo='mx', geo_name='Мексика',
                                 products_num=len(mexico_prod_dict.keys()))

        """Filling in product row with formulas and values."""
        for product in mexico_prod_dict.keys():
            product_row = self.report_sheet.max_row + 1
            self.report_sheet.cell(row=product_row, column=1).value = product
            self.report_sheet.cell(row=product_row, column=2).value = mexico_prod_dict[product][0]
            self.report_sheet.cell(row=product_row, column=5).value = mexico_prod_dict[product][1]
            self.report_sheet.cell(row=product_row, column=6).value = f'=E{product_row}/B{product_row}'
            self.report_sheet.cell(row=product_row,
                                   column=8).value = mexico_prod_dict[product][2] / RATES['mx']

            self.report_sheet.cell(row=product_row, column=7).value = f'=F{product_row}-O{product_row}'
            self.report_sheet.cell(row=product_row, column=9).value = f'=H{product_row}-N{product_row}'

            mexican_targets = TARGETS.get('mx')
            mexican_product_targets = mexican_targets.get(product)
            if mexican_product_targets:
                self.report_sheet.cell(row=product_row,
                                       column=column_index_from_string('N')).fill = yellow_fill
                self.report_sheet.cell(row=product_row,
                                       column=column_index_from_string('N')).value = mexican_product_targets[0]
                self.report_sheet.cell(row=product_row,
                                       column=column_index_from_string('O')).fill = yellow_fill
                self.report_sheet.cell(row=product_row,
                                       column=column_index_from_string('O')).value = mexican_product_targets[1] / 100
                self.report_sheet.cell(row=product_row,
                                       column=column_index_from_string('O')).number_format = '0%'

        self.report_sheet.row_dimensions.group(country_row + 1, country_row + len(mexico_prod_dict.keys()), hidden=True)

    def download_geo_data_and_report_compilation(self):
        for geo in self.geo_dict.keys():
            country_row = self.report_sheet.max_row + 1
            if geo == 'mx':
                self.compiling_with_mex_info(country_row=country_row)
            else:

                geo_name = self.geo_dict[geo][0]
                info_by_geo = requests.get(
                    f'{os.getenv("platform_offers_api_link")}&from={self.start_date}&to={self.end_date}&geo={geo}')
                products_num = len(info_by_geo.json())
                self.country_row_fill_in(country_row=country_row, geo=geo, geo_name=geo_name, products_num=products_num)
                self.product_row_fill_in(info_by_geo=info_by_geo, geo=geo, country_row=country_row,
                                         products_num=products_num)
            self.country_rows.append(country_row)

    def total_row_fill_in(self):

        self.total_row = self.report_sheet.max_row + 1
        grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        for column in range(1, column_index_from_string('L') + 1):
            self.report_sheet.cell(row=self.total_row, column=column).fill = grey_fill

        self.report_sheet.cell(row=self.total_row, column=column_index_from_string('A')).value = 'Всего'

        """Columns B, D, E fill - sum of the country rows values."""
        total_B = f'=B{self.country_rows[0]}'
        total_D = f'=D{self.country_rows[0]}'
        total_E = f'=E{self.country_rows[0]}'

        for country_row in self.country_rows[1:]:
            total_B += f'+B{country_row}'
            total_D += f'+D{country_row}'
            total_E += f'+E{country_row}'

        self.report_sheet.cell(row=self.total_row, column=column_index_from_string('B')).value = total_B
        self.report_sheet.cell(row=self.total_row, column=column_index_from_string('D')).value = total_D
        self.report_sheet.cell(row=self.total_row, column=column_index_from_string('E')).value = total_E
        self.report_sheet.cell(row=self.total_row,
                               column=column_index_from_string('F')).value = f'=E{self.total_row}/B{self.total_row}'
        self.report_sheet.cell(row=self.total_row, column=column_index_from_string('F')).number_format = '0%'
        self.report_sheet.cell(row=self.total_row,
                               column=column_index_from_string('J')).value = f'=SUM(J3:J{self.total_row - 1})'
        self.report_sheet.cell(row=self.total_row,
                               column=column_index_from_string('K')).value = f'=SUM(K3:K{self.total_row - 1})'
        self.report_sheet.cell(row=self.total_row,
                               column=column_index_from_string('L')).value = f'=SUM(L3:L{self.total_row - 1})'

    def delivery_rate_tables_compile(self):
        report_date = self.start_date.split('-')
        report_day = int(report_date[2])
        report_month = int(report_date[1])
        report_year = int(report_date[0])
        report_date_datetime_format = datetime.datetime(year=report_year, month=report_month, day=report_day)
        two_months_ago_number = (report_date_datetime_format - relativedelta(months=2)).month
        two_months_ago_year = (report_date_datetime_format - relativedelta(months=2)).year
        one_month_ago_number = (report_date_datetime_format - relativedelta(months=1)).month
        one_month_ago_year = (report_date_datetime_format - relativedelta(months=1)).year
        delivery_table_row = self.total_row + 3
        delivery_table_1_columns = (1, 6)
        delivery_table_2_columns = (8, 12)

        delivery_table_row_names = {}  # row names in format {'geo name'(str): 'row header'str()}

        """Headers"""

        if report_day <= 22:
            self.report_sheet.cell(row=delivery_table_row, column=1).value = 'Выкуп траф, %'
            self.report_sheet.cell(row=delivery_table_row, column=3).value = MONTHS[two_months_ago_number]
            self.report_sheet.cell(row=delivery_table_row + 1, column=1).value = 'Дата'
            self.report_sheet.cell(row=delivery_table_row + 1, column=2).value = '1-7'
            self.report_sheet.cell(row=delivery_table_row + 1, column=3).value = '8-15'
            self.report_sheet.cell(row=delivery_table_row + 1, column=4).value = '16-22'
            self.report_sheet.cell(row=delivery_table_row + 1,
                                   column=6).value = f'23-{calendar.monthrange(two_months_ago_year, two_months_ago_number)[1]}'

            self.report_sheet.cell(row=delivery_table_row, column=8).value = 'Выкуп траф, %'
            self.report_sheet.cell(row=delivery_table_row, column=10).value = MONTHS[one_month_ago_number]
            self.report_sheet.cell(row=delivery_table_row + 1, column=8).value = 'Дата'
            self.report_sheet.cell(row=delivery_table_row + 1, column=9).value = '1-7'
            self.report_sheet.cell(row=delivery_table_row + 1, column=10).value = '8-15'
            self.report_sheet.cell(row=delivery_table_row + 1, column=11).value = '16-22'
            self.report_sheet.cell(row=delivery_table_row + 1,
                                   column=12).value = f'23-{calendar.monthrange(one_month_ago_year, one_month_ago_number)[1]}'

        else:
            self.report_sheet.cell(row=delivery_table_row, column=1).value = 'Выкуп траф, %'
            self.report_sheet.cell(row=delivery_table_row, column=3).value = MONTHS[one_month_ago_number]
            self.report_sheet.cell(row=delivery_table_row + 1, column=1).value = 'Дата'
            self.report_sheet.cell(row=delivery_table_row + 1, column=2).value = '1-7'
            self.report_sheet.cell(row=delivery_table_row + 1, column=3).value = '8-15'
            self.report_sheet.cell(row=delivery_table_row + 1, column=4).value = '16-22'
            self.report_sheet.cell(row=delivery_table_row + 1,
                                   column=6).value = f'23-{calendar.monthrange(one_month_ago_year, one_month_ago_number)[1]}'

            self.report_sheet.cell(row=delivery_table_row, column=8).value = 'Выкуп траф, %'
            self.report_sheet.cell(row=delivery_table_row, column=10).value = MONTHS[report_month]
            self.report_sheet.cell(row=delivery_table_row + 1, column=8).value = 'Дата'
            self.report_sheet.cell(row=delivery_table_row + 1, column=9).value = '1-7'
            self.report_sheet.cell(row=delivery_table_row + 1, column=10).value = '8-15'
            self.report_sheet.cell(row=delivery_table_row + 1, column=11).value = '16-22'
            self.report_sheet.cell(row=delivery_table_row + 1,
                                   column=12).value = f'23-{calendar.monthrange(report_year, report_month)[1]}'

        for geo in self.geo_dict.keys():
            if report_day < 23:
                table_1_month = two_months_ago_number
                table_1_year = two_months_ago_year
                table_2_month = one_month_ago_number
                table_2_year = one_month_ago_year
            else:
                table_1_month = one_month_ago_number
                table_1_year = one_month_ago_year
                table_2_month = report_month
                table_2_year = report_year
            geo_row = self.report_sheet.max_row + 1
            self.report_sheet.cell(row=geo_row, column=1).value = delivery_table_row_names.get(geo)
            self.report_sheet.cell(row=geo_row, column=8).value = delivery_table_row_names.get(geo)
            self.report_sheet.cell(row=geo_row, column=2).value = delivery_rates(geo=geo,
                                                                                 start_date=f'{table_1_year}-{table_1_month}-01',
                                                                                 end_date=f'{table_1_year}-{table_1_month}-07')

            self.report_sheet.cell(row=geo_row, column=3).value = delivery_rates(geo=geo,
                                                                                 start_date=f'{table_1_year}-{table_1_month}-08',
                                                                                 end_date=f'{table_1_year}-{table_1_month}-15')
            self.report_sheet.cell(row=geo_row, column=4).value = delivery_rates(geo=geo,
                                                                                 start_date=f'{table_1_year}-{table_1_month}-16',
                                                                                 end_date=f'{table_1_year}-{table_1_month}-22')
            self.report_sheet.cell(row=geo_row, column=6).value = delivery_rates(geo=geo,
                                                                                 start_date=f'{table_1_year}-{table_1_month}-23',
                                                                                 end_date=f'{table_1_year}-{table_1_month}-{calendar.monthrange(table_1_year, table_1_month)[1]}')

            self.report_sheet.cell(row=geo_row, column=9).value = delivery_rates(geo=geo,
                                                                                 start_date=f'{table_2_year}-{table_2_month}-01',
                                                                                 end_date=f'{table_2_year}-{table_2_month}-07')
            if report_day <= 7:
                self.report_sheet.cell(row=geo_row, column=10).value = delivery_rates(geo=geo,
                                                                                      start_date=f'{table_2_year}-{table_2_month}-08',
                                                                                      end_date=f'{table_2_year}-{table_2_month}-15')

            if 8 <= report_day <= 22:
                self.report_sheet.cell(row=geo_row, column=10).value = delivery_rates(geo=geo,
                                                                                      start_date=f'{table_2_year}-{table_2_month}-08',
                                                                                      end_date=f'{table_2_year}-{table_2_month}-15')
                self.report_sheet.cell(row=geo_row, column=11).value = delivery_rates(geo=geo,
                                                                                      start_date=f'{table_2_year}-{table_2_month}-16',
                                                                                      end_date=f'{table_2_year}-{table_2_month}-22')

            if 16 <= report_day <= 22:
                self.report_sheet.cell(row=geo_row, column=12).value = delivery_rates(geo=geo,
                                                                                      start_date=f'{table_2_year}-{table_2_month}-23',
                                                                                      end_date=f'{table_2_year}-{table_2_month}-{calendar.monthrange(table_2_year, table_2_month)[1]}')

        """Setting fonts borders, color."""
        bold = Font(bold=True)
        grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        Border(left=Side(style='thin'),
               bottom=Side(style='thin'),
               right=Side(style='thin'),
               top=Side(style='thin'))

        def delivery_tables_styling(col_start, col_finish):
            for column_index in range(col_start, col_finish + 1):
                self.report_sheet.cell(row=delivery_table_row, column=column_index).font = bold
                self.report_sheet.cell(row=delivery_table_row, column=column_index).fill = grey_fill
                self.report_sheet.cell(row=delivery_table_row, column=column_index).border = Border(
                    bottom=Side(style='thin'),
                    top=Side(style='thin'))

                self.report_sheet.cell(row=delivery_table_row + 1, column=column_index).font = bold
                self.report_sheet.cell(row=delivery_table_row + 1, column=column_index).fill = grey_fill

                for row in range(delivery_table_row + 1, self.report_sheet.max_row + 1):
                    self.report_sheet.cell(row=row, column=column_index).border = Border(
                        bottom=Side(style='thin'),
                        top=Side(style='thin'),
                        right=Side(style='thin'),
                        left=Side(style='thin'))

                    if column_index == 1 or column_index == 8:
                        self.report_sheet.cell(row=row, column=column_index).font = bold
            self.report_sheet.cell(row=delivery_table_row, column=6).border = Border(
                bottom=Side(style='thin'),
                top=Side(style='thin'),
                right=Side(style='thin'))
            self.report_sheet.cell(row=delivery_table_row, column=8).border = Border(
                bottom=Side(style='thin'),
                top=Side(style='thin'),
                left=Side(style='thin'))
            self.report_sheet.cell(row=delivery_table_row, column=12).border = Border(
                bottom=Side(style='thin'),
                top=Side(style='thin'),
                right=Side(style='thin'))

        delivery_tables_styling(delivery_table_1_columns[0], delivery_table_1_columns[1])
        delivery_tables_styling(delivery_table_2_columns[0], delivery_table_2_columns[1])

    def setting_columns_format_and_freezing_panes(self):
        self.report_sheet.freeze_panes = self.report_sheet.cell(row=3, column=2)

        """Setting percentage format to columns C, F, G, H, I"""
        for column_letter in ['C', 'F', 'G']:
            for row in range(3, self.report_sheet.max_row + 1):
                self.report_sheet.cell(row=row, column=column_index_from_string(column_letter)).number_format = '0%'

        for column_letter in ['H', 'I']:
            for row in range(3, self.report_sheet.max_row + 1):
                self.report_sheet.cell(row=row, column=column_index_from_string(column_letter)).number_format = '0'

        """Setting borders to main table"""
        main_table_border = Border(left=Side(style='thin'),
                                   bottom=Side(style='thin'),
                                   right=Side(style='thin'),
                                   top=Side(style='thin'))
        for column in range(1, column_index_from_string('L') + 1):
            for row in range(3, self.report_sheet.max_row + 1):
                self.report_sheet.cell(row=row, column=column).border = main_table_border

        """Hiding columns E, N, O."""
        for column in ['E', 'N', 'O']:
            self.report_sheet.column_dimensions[column].hidden = True

    def saving_xlsx_file(self):
        save_file = input('enter file address to be saved\n')
        self.opened_xlsx_file.save(save_file)
        self.opened_xlsx_file.close()

    def action(self):
        self.fill_geo_dict()
        self.connect_with_xlsx_file_and_sheet_create()
        self.sheet_headers_filling_in_and_columns_dimensions()
        self.download_geo_data_and_report_compilation()
        self.total_row_fill_in()
        self.setting_columns_format_and_freezing_panes()
        self.delivery_rate_tables_compile()
        self.saving_xlsx_file()


date = datetime.datetime.today() - datetime.timedelta(days=3)
a = StatisticsProcessor(f'{date.date()}')
a.action()
